from openpyxl import Workbook
import re

rownumber=1 #Row number of the excel file where each row links to one paper/entity (each subject of the NT file).
visitedsubjects = [] #List of subjects already visited
NTfilepath = "ISWC.nt" #Input NT file
XlsxPath = "metis.xlsx" #Excel File to which we have for output.

wb = Workbook() #open workbook
ws = wb.active # select the active worksheet# select the active worksheet

#Set up Header Row. Contents of Row 1 of the Excel File.
ws.cell(row=1, column=1, value= 'Node Number')
ws.cell(row=1, column=2, value= 'Title')
ws.cell(row=1, column=3, value= 'Number of Authors')
ws.cell(row=1, column=4, value= 'Year')
ws.cell(row=1, column=5, value= 'Same As')
ws.cell(row=1, column=6, value='Published in Book')

def checkpredicate(predicate): #Check if we need the predicate and if we do then return to which column we should save it.
    if(predicate == "http://dblp.org/rdf/schema-2017-04-18#title"):
        return 2
    if (predicate =="http://dblp.org/rdf/schema-2017-04-18#authoredBy"):
        return 3
    if (predicate =="http://dblp.org/rdf/schema-2017-04-18#yearOfPublication"):
        return 4
    if (predicate =="http://www.w3.org/2002/07/owl#sameAs"):
        return 5
    if (predicate =="http://dblp.org/rdf/schema-2017-04-18#publishedAsPartOf"):
        return 6
    print "\nUnrequired Predicate Found" + predicate
    d.write("\nUnrequired Predicate Found: %s " %predicate) #Save the predicates not considered to cross-check.
    return 10

def checksubject(subject): #All triples for the same paper will have the same subject and
                           #belong in the same row. If subject changes we go to the next row.
    global rownumber
    global visitedsubjects
    if (subject not in visitedsubjects): # A new paper
        rownumber+=1                    #Go to the next row on the Excel File
        visitedsubjects.append(subject) #Append the current subject to the list of visited subjects.
        ws.cell(row=rownumber, column=1, value=rownumber-1) #Column 1 of excel file contains conf. paper number
        ws.cell(row=rownumber, column=3, value=0)           #Column 3 of excel file is number of authors. Set first to zero.
    return

#Execution starts from here:-
with open(NTfilepath) as f, open("deleted predicates.txt", 'w+') as d: #open input NT file and another file to save the predicates ignored.
        for line in f:                                  #Iterate through the NT file line by line
            eachnumber= re.findall("\<(.*?)\>", line)   #Get the subject, predicate and object stored in eachnumber
            eachnumber[0]=eachnumber[0].strip("<>.")    #Subject
            eachnumber[1]=eachnumber[1].strip("<>.")    #Predicate
            eachnumber[2]=eachnumber[2].strip("<>.")    #Object
            checksubject(eachnumber[0])     #Set the row number to write on the Excel file depending on the subject.
            columnnumber= checkpredicate(eachnumber[1]) #Get column number of the excel file where to be saved for the predicate of the current triple.

            if (columnnumber == 3): #Number of authors column. We add the names of the authors at the right end columns 7 onwards
                temp = ws.cell(row = rownumber, column= 3).value #Get the current "number of author" column value of the excel file to know which column to write the name of the author.
                authorcol=temp+7 #Get the free column number of the excel file to write the name of author.
                ws.cell(row=rownumber, column=authorcol, value=eachnumber[2]) #Write the name of author.
                ws.cell(row=rownumber, column=3, value=temp+1) #Update the "number of authors" column of the excel file.

            if (columnnumber<10 and columnnumber!=3): #Predicate is to be considered but value is not "author"
                ws.cell(row=rownumber,column=columnnumber,value=eachnumber[2]) #Save the predicate to the particular cell of the excel file.

        f.close()
        d.close()
        wb.save(XlsxPath)
